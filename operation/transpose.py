import openpyxl

from datetime import datetime, timedelta


class Excel(object):

    def __init__(self, file_name):
        self.file_name = file_name
        self.data = openpyxl.load_workbook(self.file_name)
        self.sheet_names = self.data.sheetnames
        print('data sheet name is %s' % self.sheet_names)

    def print_excel_info(self):
        print('file name is %s' % self.file_name)

    def read(self, sheet_name):
        sheet_content = self.data[sheet_name]

        new_content = []
        new_content.append(['Employee / Vendor / Client Name', 'Date', 'Task Name', 'Quantity', 'Project Description'])
        for i in range(2, sheet_content.max_row + 1):
            # for j in range(1, sheet_content.max_column + 1):
            employee_name = sheet_content.cell(row=i, column=2).value
            start_column = 7
            for j in range(5):
                quantity = sheet_content.cell(row=i, column=start_column + j).value
                if quantity == None or 0 == quantity or (str(quantity)).isspace():
                    print("0 quantity")
                    pass
                else:
                    week_day = self.get_week_n_data(j)
                    task_name = sheet_content.cell(row=i, column=12).value
                    project_description = sheet_content.cell(row=i, column=13).value
                    new_row = [employee_name, week_day, task_name, quantity, project_description]
                    new_content.append(new_row)

        self.new_content = new_content

    def get_week_n_data(self, d):
        theDay = datetime.today()
        theDay_before = theDay - timedelta(days=7)
        weekMonday = theDay_before - timedelta(days=theDay_before.weekday() - d)
        result = weekMonday.strftime('%Y-%m-%d')
        return result

    def write(self, sheet_name):
        self.data.create_sheet(title=sheet_name)
        self.data.save(self.file_name)

        self.sheet_names = self.data.sheetnames
        new_sheet = self.data[sheet_name]

        for row in self.new_content:
            new_sheet.append(row)
        self.data.save(self.file_name)
